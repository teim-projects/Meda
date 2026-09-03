import io
import openpyxl
from datetime import date
from decimal import Decimal
from django.test import TestCase, Client
from django.contrib.auth.models import User
from django.urls import reverse

from .models import Biomass, Bagasse, MSW, SHP, GovtSolarization, SolarGrid, SolarKusum, Wind

class EnergyTemplatesTests(TestCase):
    def setUp(self):
        # Create user and log in client
        self.user = User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='adminpassword'
        )
        self.client = Client()
        self.client.force_login(self.user)
        
        # Test configurations for all 8 active types
        self.active_types = ['biomass', 'bagasse', 'govt-solarization', 'msw', 'shp', 'solar-grid', 'solar-kusum', 'wind']

    def test_download_templates(self):
        """
        Verify that blank template download works and contains the correct headers for all types.
        """
        for etype in self.active_types:
            url = reverse('energy-template-download', kwargs={'energy_type': etype})
            response = self.client.get(url)
            self.assertEqual(response.status_code, 200, f"Failed to download template for {etype}")
            self.assertEqual(
                response['Content-Type'],
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # Read the response spreadsheet to check headers
            wb = openpyxl.load_workbook(io.BytesIO(response.content))
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows())]
            self.assertTrue(len(headers) > 0)

    def test_export_data(self):
        """
        Verify that data export works for all types.
        """
        for etype in self.active_types:
            url = reverse('energy-template-export', kwargs={'energy_type': etype})
            response = self.client.get(url)
            self.assertEqual(response.status_code, 200, f"Failed to export data for {etype}")
            self.assertEqual(
                response['Content-Type'],
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

    def test_upload_excel_govt_solarization(self):
        """
        Verify that uploading a valid Excel sheet for Government Building Solarization successfully creates a record.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        
        headers = ['Department', 'Division', 'District', 'Taluka', 'Consumer No', 'Building Name', 'Address', 'PIN Code', 'Ownership Type', 'Capacity (kW)', 'Capacity (MW)']
        ws.append(headers)
        
        row = ['Public Works Dept', 'West Div', 'Pune', 'Haveli', 'CONS123456', 'PWD Main Building', '123 Station Road', '411001', 'Government Owned', 50.5, 0.050]
        ws.append(row)
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'govt_solarization_test.xlsx'
        
        url = reverse('energy-template-upload', kwargs={'energy_type': 'govt-solarization'})
        response = self.client.post(url, {'file': excel_file})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        
        obj = GovtSolarization.objects.first()
        self.assertIsNotNone(obj)
        self.assertEqual(obj.department, 'Public Works Dept')
        self.assertEqual(obj.building_name, 'PWD Main Building')
        self.assertEqual(obj.capacity_kw, Decimal('50.500'))

    def test_upload_excel_biomass(self):
        """
        Verify that uploading a valid Excel sheet for Biomass successfully creates a record with updated headers.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        
        headers = ['Name of Developer', 'Site Name', 'Source', 'Taluka', 'Village', 'Investor Name & Address', 'District', 'Capacity MW', 'Commission date']
        ws.append(headers)
        
        row = ['Bio Energy Developer', 'Site One', 'Bio Agri', 'Haveli', 'Hadapsar', 'Investor Biomass Co', 'Pune', 12.5, '2024-05-20']
        ws.append(row)
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'biomass_test.xlsx'
        
        url = reverse('energy-template-upload', kwargs={'energy_type': 'biomass'})
        response = self.client.post(url, {'file': excel_file})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        
        obj = Biomass.objects.first()
        self.assertIsNotNone(obj)
        self.assertEqual(obj.name_of_developer, 'Bio Energy Developer')
        self.assertEqual(obj.site_name, 'Site One')
        self.assertEqual(obj.source, 'Bio Agri')
        self.assertEqual(obj.taluka, 'Haveli')
        self.assertEqual(obj.village, 'Hadapsar')
        self.assertEqual(obj.investor_name_and_address, 'Investor Biomass Co')
        self.assertEqual(obj.district, 'Pune')
        self.assertEqual(obj.capacity_mw, Decimal('12.500'))
        self.assertEqual(str(obj.commissioned_date), '2024-05-20')

    def test_upload_excel_bagasse(self):
        """
        Verify that uploading a valid Excel sheet for Bagasse successfully creates a record.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        
        headers = ['Source', 'Name of Developer', 'Site Name', 'Investor Name & Address', 'Village', 'Taluka', 'District', 'Capacity MW', 'Commissioned Date']
        ws.append(headers)
        
        row = ['Sugar Mill', 'Dev Bagasse', 'Site Alpha', 'Investor Address XYZ', 'Vil B', 'Tal B', 'Dist B', 15.75, '2024-03-15']
        ws.append(row)
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'bagasse_test.xlsx'
        
        url = reverse('energy-template-upload', kwargs={'energy_type': 'bagasse'})
        response = self.client.post(url, {'file': excel_file})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        
        obj = Bagasse.objects.first()
        self.assertIsNotNone(obj)
        self.assertEqual(obj.source, 'Sugar Mill')
        self.assertEqual(obj.name_of_developer, 'Dev Bagasse')
        self.assertEqual(obj.capacity_mw, Decimal('15.750'))

    def test_upload_excel_msw(self):
        """
        Verify that uploading a valid Excel sheet for MSW successfully creates a record.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        
        headers = ['Source', 'Investor Name & Address', 'Name of Developer', 'Site Name', 'Village', 'Taluka', 'District', 'Capacity MW', 'Commissioned Date', 'Grid Type']
        ws.append(headers)
        
        row = ['Municipal Waste', 'Investor MSW', 'MSW Developer', 'Waste Plant 1', 'Vil C', 'Tal C', 'Dist C', 8.5, '2024-04-10', 'Grid Connected']
        ws.append(row)
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'msw_test.xlsx'
        
        url = reverse('energy-template-upload', kwargs={'energy_type': 'msw'})
        response = self.client.post(url, {'file': excel_file})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        
        obj = MSW.objects.first()
        self.assertIsNotNone(obj)
        self.assertEqual(obj.grid_type, 'Grid Connected')
        self.assertEqual(obj.capacity_mw, Decimal('8.500'))

    def test_upload_excel_shp(self):
        """
        Verify that uploading a valid Excel sheet for SHP successfully creates a record.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        
        headers = ['Completed Hydro Electric Projects', 'Date of Commissioning', 'Region', 'Village/Taluka', 'District', 'Capacity (MW)', 'Comissioning Year']
        ws.append(headers)
        
        row = ['Hydro Project One', '2024-06-20', 'Western Region', 'Vil/Tal A', 'District X', 5.67, '2024-25']
        ws.append(row)
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'shp_test.xlsx'
        
        url = reverse('energy-template-upload', kwargs={'energy_type': 'shp'})
        response = self.client.post(url, {'file': excel_file})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        
        obj = SHP.objects.first()
        self.assertIsNotNone(obj)
        self.assertEqual(obj.completed_hydro_electric_projects, 'Hydro Project One')
        self.assertEqual(obj.region, 'Western Region')
        self.assertEqual(obj.district, 'District X')
        self.assertEqual(obj.installed_capacity_mw, Decimal('5.670'))

    def test_upload_excel_reordered_columns(self):
        """
        Verify that uploading an Excel file with columns arranged in a different order accurately maps values to proper database fields.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Columns in completely scrambled / reordered sequence
        shuffled_headers = ['Capacity MW', 'District', 'Source', 'Commissioned Date', 'Name of Developer', 'Taluka', 'Village', 'Site Name', 'Investor Name & Address']
        ws.append(shuffled_headers)
        
        shuffled_row = [24.50, 'Satara', 'Biomass Source', '2024-05-12', 'Developer Shuffled', 'Koregaon', 'Pimpri', 'Site Beta', 'Investor Addr ABC']
        ws.append(shuffled_row)
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'bagasse_reordered.xlsx'
        
        url = reverse('energy-template-upload', kwargs={'energy_type': 'bagasse'})
        response = self.client.post(url, {'file': excel_file})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        
        # Verify db insert to exact fields despite scrambled column positions
        obj = Bagasse.objects.filter(name_of_developer='Developer Shuffled').first()
        self.assertIsNotNone(obj)
        self.assertEqual(obj.capacity_mw, Decimal('24.500'))
        self.assertEqual(obj.district, 'Satara')
        self.assertEqual(obj.source, 'Biomass Source')
        self.assertEqual(obj.name_of_developer, 'Developer Shuffled')
        self.assertEqual(obj.taluka, 'Koregaon')
        self.assertEqual(obj.village, 'Pimpri')
        self.assertEqual(obj.site_name, 'Site Beta')
        self.assertEqual(obj.investor_name_and_address, 'Investor Addr ABC')

    def test_upload_excel_solar_kusum(self):
        """
        Verify that uploading a valid Excel sheet for Solar Kusum successfully creates a record with exact specified columns:
        Type, Developer Name, Project Location, Commissioned Capacity (MW), Commission Date, district.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        
        headers = ['Type', 'Developer Name', 'Project Location', 'Commissioned Capacity (MW)', 'Commission Date', 'district']
        ws.append(headers)
        
        row = ['Component A', 'SunPower Ltd', 'Talwade Block 4', 2.5, '2024-06-18', 'Nashik']
        ws.append(row)
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'solar_kusum_test.xlsx'
        
        url = reverse('energy-template-upload', kwargs={'energy_type': 'solar-kusum'})
        response = self.client.post(url, {'file': excel_file})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        
        obj = SolarKusum.objects.first()
        self.assertIsNotNone(obj)
        self.assertEqual(obj.type, 'Component A')
        self.assertEqual(obj.developer_name, 'SunPower Ltd')
        self.assertEqual(obj.project_location, 'Talwade Block 4')
        self.assertEqual(obj.commissioned_capacity_mw, Decimal('2.500'))
        self.assertEqual(str(obj.commission_date), '2024-06-18')
        self.assertEqual(obj.district, 'Nashik')

    def test_upload_excel_solar_grid(self):
        """
        Verify that uploading a valid Excel sheet for Solar Grid successfully creates a record with exact specified columns:
        Developer Name, Project Location, District, Commissioned Capacity (MW), Commission Date.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        
        headers = ['Developer Name', 'Project Location', 'District', 'Commissioned Capacity (MW)', 'Commission Date']
        ws.append(headers)
        
        row = ['Solar Tech India', 'Village Shindewadi', 'Pune', 18.25, '2024-04-10']
        ws.append(row)
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'solar_grid_test.xlsx'
        
        url = reverse('energy-template-upload', kwargs={'energy_type': 'solar-grid'})
        response = self.client.post(url, {'file': excel_file})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        
        obj = SolarGrid.objects.first()
        self.assertIsNotNone(obj)
        self.assertEqual(obj.developer_name, 'Solar Tech India')
        self.assertEqual(obj.project_location, 'Village Shindewadi')
        self.assertEqual(obj.district, 'Pune')
        self.assertEqual(obj.commissioned_capacity_mw, Decimal('18.250'))
        self.assertEqual(str(obj.commission_date), '2024-04-10')

    def test_energy_analytics_govt_solarization(self):
        """
        Verify that energy-analytics endpoint returns total_projects, total_capacity_mw, and district breakdown.
        """
        GovtSolarization.objects.create(
            department='Forestry',
            division='Nagpur Div',
            district='NAGPUR',
            taluka='Nagpur Urban',
            building_name='Forest Office',
            capacity_kw=Decimal('100.000'),
            capacity_mw=Decimal('0.100')
        )
        GovtSolarization.objects.create(
            department='Health',
            division='Pune Div',
            district='PUNE',
            taluka='Haveli',
            building_name='District Hospital',
            capacity_kw=Decimal('200.000'),
            capacity_mw=Decimal('0.200')
        )

        url = reverse('energy-analytics', kwargs={'energy_type': 'govt-solarization'})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['success'])
        self.assertEqual(data['total_projects'], 2)
        self.assertEqual(data['total_capacity_mw'], 0.3)
        self.assertEqual(len(data['districts']), 2)
        self.assertEqual(data['districts'][0]['district'], 'Nagpur')

    def test_wind_exact_columns_and_upload(self):
        """
        Verify that Wind template has exact columns:
        Developer, Investor, Capacity MW, Standardized Date, Date of Commissioned, Gut No., Taluka, Village, District, Site Name, Source, Year
        and that Excel upload saves data correctly.
        """
        # 1. Verify headers from energy-data-list endpoint
        url_headers = reverse('energy-data-list', kwargs={'energy_type': 'wind'})
        res_headers = self.client.get(url_headers).json()
        expected = [
            'Developer', 'Investor', 'Capacity MW', 'Standardized Date',
            'Date of Commissioned', 'Gut No.', 'Taluka', 'Village',
            'District', 'Site Name', 'Source', 'Year'
        ]
        self.assertEqual(res_headers['headers'], expected)

        # 2. Upload Excel with exact columns
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(expected)
        long_investor = "Tata Power Renewable Energy Limited Consortium " * 8  # ~376 chars
        ws.append([
            'Suzlon Energy', long_investor, 2.1, '2023-04-15',
            '2023-04-15', 'Gut 102/1', 'Khatav', 'Vardhangad',
            'Satara', 'Vardhangad Wind Farm', 'Wind Power', '2023'
        ])

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'wind_test.xlsx'

        url_upload = reverse('energy-template-upload', kwargs={'energy_type': 'wind'})
        res_upload = self.client.post(url_upload, {'file': excel_file})
        self.assertEqual(res_upload.status_code, 200)
        self.assertTrue(res_upload.json()['success'])

        wind_obj = Wind.objects.first()
        self.assertIsNotNone(wind_obj)
        self.assertEqual(wind_obj.developer, 'Suzlon Energy')
        self.assertEqual(wind_obj.investor, long_investor.strip())
        self.assertEqual(wind_obj.capacity_mw, Decimal('2.100'))
        self.assertEqual(str(wind_obj.standardized_date), '2023-04-15')
        self.assertEqual(str(wind_obj.date_of_commissioned), '2023-04-15')
        self.assertEqual(wind_obj.gut_no, 'Gut 102/1')
        self.assertEqual(wind_obj.taluka, 'Khatav')
        self.assertEqual(wind_obj.village, 'Vardhangad')
        self.assertEqual(wind_obj.district, 'Satara')
        self.assertEqual(wind_obj.site_name, 'Vardhangad Wind Farm')
        self.assertEqual(wind_obj.source, 'Wind Power')
        self.assertEqual(wind_obj.year, '2023')


    def test_energy_analytics_msw_shp_kusum(self):
        """
        Verify analytics endpoints for MSW, SHP, and Solar Kusum.
        """
        MSW.objects.create(
            district='Pune',
            capacity_mw=Decimal('14.000'),
            name_of_developer='Antony Lara',
            site_name='Moshi Pune'
        )
        url_msw = reverse('energy-analytics', kwargs={'energy_type': 'msw'})
        res_msw = self.client.get(url_msw).json()
        self.assertTrue(res_msw['success'])
        self.assertEqual(res_msw['total_projects'], 1)
        self.assertEqual(res_msw['total_capacity_mw'], 14.0)

        SHP.objects.create(
            district='Kolhapur',
            installed_capacity_mw=Decimal('3.500'),
            completed_hydro_electric_projects='Ambai Hep',
            comissioning_year='2012'
        )
        url_shp = reverse('energy-analytics', kwargs={'energy_type': 'shp'})
        res_shp = self.client.get(url_shp).json()
        self.assertTrue(res_shp['success'])
        self.assertEqual(res_shp['total_projects'], 1)
        self.assertEqual(res_shp['total_capacity_mw'], 3.5)

        SolarKusum.objects.create(
            type='KUSUM A',
            developer_name='ABC Power',
            district='Ahmednagar',
            commissioned_capacity_mw=Decimal('2.000'),
            commission_date='2024-05-10'
        )
        url_kusum = reverse('energy-analytics', kwargs={'energy_type': 'solar-kusum'})
        res_kusum = self.client.get(url_kusum).json()
        self.assertTrue(res_kusum['success'])
        self.assertEqual(res_kusum['total_projects'], 1)
        self.assertEqual(res_kusum['types']['KUSUM A']['capacity_mw'], 2.0)






