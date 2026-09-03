import io
import datetime
import openpyxl
from decimal import Decimal
from django.db import models
from django.http import HttpResponse
from rest_framework import status, permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser

from .models import Biomass, Bagasse, MSW, SHP, GovtSolarization, SolarGrid, SolarKusum, Wind, MSKVY

# Dynamic maps for energy types, including model references and exact header mapping
CONFIG_MAP = {
    'biomass': {
        'model': Biomass,
        'display_name': 'Biomass',
        'headers': [
            'Name of Developer',
            'Site Name',
            'Source',
            'Taluka',
            'Village',
            'Investor Name & Address',
            'District',
            'Capacity MW',
            'Commission date'
        ],
        'fields': {
            'Name of Developer': 'name_of_developer',
            'Site Name': 'site_name',
            'Source': 'source',
            'Taluka': 'taluka',
            'Village': 'village',
            'Investor Name & Address': 'investor_name_and_address',
            'District': 'district',
            'Capacity MW': 'capacity_mw',
            'Commission date': 'commissioned_date',
            'Commissioned Date': 'commissioned_date'
        }
    },
    'bagasse': {
        'model': Bagasse,
        'display_name': 'Bagasse',
        'headers': ['Source', 'Name of Developer', 'Site Name', 'Investor Name & Address', 'Village', 'Taluka', 'District', 'Capacity MW', 'Commissioned Date'],
        'fields': {
            'Source': 'source',
            'Name of Developer': 'name_of_developer',
            'Site Name': 'site_name',
            'Investor Name & Address': 'investor_name_and_address',
            'Village': 'village',
            'Taluka': 'taluka',
            'District': 'district',
            'Capacity MW': 'capacity_mw',
            'Commissioned Date': 'commissioned_date'
        }
    },
    'govt_solarization': {
        'model': GovtSolarization,
        'display_name': 'Government Building Solarization',
        'headers': ['Department', 'Division', 'District', 'Taluka', 'Consumer No', 'Building Name', 'Address', 'PIN Code', 'Ownership Type', 'Capacity (kW)', 'Capacity (MW)'],
        'fields': {
            'Department': 'department',
            'Division': 'division',
            'District': 'district',
            'Taluka': 'taluka',
            'Consumer No': 'consumer_no',
            'Building Name': 'building_name',
            'Address': 'address',
            'PIN Code': 'pin_code',
            'Ownership Type': 'ownership_type',
            'Capacity (kW)': 'capacity_kw',
            'Capacity (MW)': 'capacity_mw'
        }
    },
    'government_building_solarization': {
        'model': GovtSolarization,
        'display_name': 'Government Building Solarization',
        'headers': ['Department', 'Division', 'District', 'Taluka', 'Consumer No', 'Building Name', 'Address', 'PIN Code', 'Ownership Type', 'Capacity (kW)', 'Capacity (MW)'],
        'fields': {
            'Department': 'department',
            'Division': 'division',
            'District': 'district',
            'Taluka': 'taluka',
            'Consumer No': 'consumer_no',
            'Building Name': 'building_name',
            'Address': 'address',
            'PIN Code': 'pin_code',
            'Ownership Type': 'ownership_type',
            'Capacity (kW)': 'capacity_kw',
            'Capacity (MW)': 'capacity_mw'
        }
    },
    'msw': {
        'model': MSW,
        'display_name': 'MSW (Municipal Solid Waste)',
        'headers': ['Source', 'Investor Name & Address', 'Name of Developer', 'Site Name', 'Village', 'Taluka', 'District', 'Capacity MW', 'Commissioned Date', 'Grid Type'],
        'fields': {
            'Source': 'source',
            'Investor Name & Address': 'investor_name_and_address',
            'Name of Developer': 'name_of_developer',
            'Site Name': 'site_name',
            'Village': 'village',
            'Taluka': 'taluka',
            'District': 'district',
            'Capacity MW': 'capacity_mw',
            'Commissioned Date': 'commissioned_date',
            'Grid Type': 'grid_type'
        }
    },
    'shp': {
        'model': SHP,
        'display_name': 'SHP (Small Hydro Power)',
        'headers': ['Completed Hydro Electric Projects', 'Date of Commissioning', 'Region', 'Village/Taluka', 'District', 'Capacity (MW)', 'Comissioning Year'],
        'fields': {
            'Completed Hydro Electric Projects': 'completed_hydro_electric_projects',
            'Date of Commissioning': 'date_of_commissioning',
            'Region': 'region',
            'Village/Taluka': 'village_taluka',
            'District': 'district',
            'Capacity (MW)': 'installed_capacity_mw',
            'Comissioning Year': 'comissioning_year'
        }
    },
    'solar_grid': {
        'model': SolarGrid,
        'display_name': 'Solar Grid',
        'headers': [
            'Developer Name',
            'Project Location',
            'District',
            'Commissioned Capacity (MW)',
            'Commission Date'
        ],
        'fields': {
            'Developer Name': 'developer_name',
            'Project Location': 'project_location',
            'District': 'district',
            'district': 'district',
            'Commissioned Capacity (MW)': 'commissioned_capacity_mw',
            'Commission Date': 'commission_date',
            'Commissioned Date': 'commission_date'
        }
    },
    'solar_kusum': {
        'model': SolarKusum,
        'display_name': 'Solar Kusum',
        'headers': [
            'Type',
            'Developer Name',
            'Project Location',
            'Commissioned Capacity (MW)',
            'Commission Date',
            'district'
        ],
        'fields': {
            'Type': 'type',
            'type': 'type',
            'Developer Name': 'developer_name',
            'Project Location': 'project_location',
            'Commissioned Capacity (MW)': 'commissioned_capacity_mw',
            'Commission Date': 'commission_date',
            'Commissioned Date': 'commission_date',
            'district': 'district',
            'District': 'district'
        }
    },
    'wind': {
        'model': Wind,
        'display_name': 'Wind',
        'headers': [
            'Developer',
            'Investor',
            'Capacity MW',
            'Standardized Date',
            'Date of Commissioned',
            'Gut No.',
            'Taluka',
            'Village',
            'District',
            'Site Name',
            'Source',
            'Year'
        ],
        'fields': {
            'Developer': 'developer',
            'developer': 'developer',
            'Investor': 'investor',
            'investor': 'investor',
            'Capacity MW': 'capacity_mw',
            'capacity_mw': 'capacity_mw',
            'Standardized Date': 'standardized_date',
            'standardized_date': 'standardized_date',
            'Date of Commissioned': 'date_of_commissioned',
            'date_of_commissioned': 'date_of_commissioned',
            'Gut No.': 'gut_no',
            'gut_no': 'gut_no',
            'Taluka': 'taluka',
            'taluka': 'taluka',
            'Village': 'village',
            'village': 'village',
            'District': 'district',
            'district': 'district',
            'Site Name': 'site_name',
            'site_name': 'site_name',
            'Source': 'source',
            'source': 'source',
            'Year': 'year',
            'year': 'year'
        }
    },
    'mskvy': {
        'model': MSKVY,
        'display_name': 'MSKVY',
        'headers': [
            'Source',
            'Project Location',
            'Commissioned Capacity (MW)',
            'Commission Date',
            'District'
        ],
        'fields': {
            'Source': 'source',
            'source': 'source',
            'Project Location': 'project_location',
            'project_location': 'project_location',
            'location': 'project_location',
            'Commissioned Capacity (MW)': 'commissioned_capacity_mw',
            'Commissioned Capacity MW': 'commissioned_capacity_mw',
            'commissioned_capacity_mw': 'commissioned_capacity_mw',
            'Capacity (MW)': 'commissioned_capacity_mw',
            'Capacity MW': 'commissioned_capacity_mw',
            'capacity_mw': 'commissioned_capacity_mw',
            'Commission Date': 'commission_date',
            'Commissioned Date': 'commission_date',
            'commission_date': 'commission_date',
            'commissioned_date': 'commission_date',
            'District': 'district',
            'district': 'district'
        }
    }
}

# Alias dictionary to allow flexible header variations in uploaded files
HEADER_ALIASES = {
    'developer': ['developer', 'developer name', 'name of developer', 'developer_name', 'name_of_developer', 'bidder'],
    'investor': ['investor', 'investor name', 'investor name & address', 'investor name and address', 'investor_name_and_address', 'investor address'],
    'standardized_date': ['standardized date', 'standardized_date', 'standard date', 'std date'],
    'type': ['type', 'project type', 'kusum type', 'category'],
    'project_location': ['project location', 'location', 'project_location', 'site location', 'site name', 'address'],
    'source': ['source', 'energy source', 'source name'],
    'name_of_developer': ['name of developer', 'developer name', 'developer', 'name_of_developer', 'developer_name'],
    'developer_name': ['developer name', 'name of developer', 'developer', 'developer_name', 'name_of_developer', 'name of successful bidder'],
    'site_name': ['site name', 'site_name', 'site', 'project site'],
    'investor_name_and_address': ['investor name & address', 'investor name and address', 'investor name', 'investor_name_and_address', 'investor address'],
    'village': ['village', 'village name'],
    'taluka': ['taluka', 'taluka name'],
    'district': ['district', 'district name'],
    'capacity_mw': [
        'capacity mw', 'capacity (mw)', 'capacity_mw', 'installed capacity (mw)', 
        'installed capacity mw', 'commissioned capacity (mw)', 'commissioned capacity mw',
        'cumulative capacity mw bio', 'cumulative capacity (mw) bio', 'capacity (in mw)'
    ],
    'installed_capacity_mw': [
        'installed capacity (mw)', 'installed capacity mw', 'capacity (mw)', 'capacity mw', 
        'capacity_mw', 'installed_capacity_mw', 'installed capacity'
    ],
    'commissioned_capacity_mw': [
        'commissioned capacity (mw)', 'commissioned capacity mw', 'commissioned capacity',
        'capacity (mw)', 'capacity mw', 'capacity_mw', 'commissioned_capacity_mw'
    ],
    'capacity_kw': [
        'capacity (kw)', 'capacity kw', 'capacity_kw', 'installed capacity (kw)', 
        'installed capacity kw', 'commissioned capacity (kw)', 'capacity (in kw)'
    ],
    'quarter': ['quarter'],
    'month': ['month'],
    'day': ['day'],
    'commissioning_year': ['commissioning year', 'commissioned year', 'commissioning_year', 'year of commissioning', 'year'],
    'comissioning_year': ['comissioning year', 'commissioning year', 'commissioned year', 'commissioning_year', 'year of commissioning', 'year'],
    'commissioned_date': ['commissioned date', 'commissioning date', 'date of commissioning', 'commission_date', 'date_of_commissioned', 'commission date'],
    'commission_date': ['commission date', 'commissioned date', 'commissioning date', 'date of commissioning', 'date_of_commissioned'],
    'date_of_commissioning': ['date of commissioning', 'commissioned date', 'commissioning date', 'commission_date', 'date_of_commissioned', 'commission date'],
    'grid_type': ['grid type', 'grid_type', 'grid connected/offgrid', 'grid connected / offgrid', 'grid connected offgrid', 'grid'],
    'grid_connected_offgrid': ['grid connected/offgrid', 'grid connected / offgrid', 'grid connected offgrid', 'grid type', 'grid_type'],
    'region': ['region', 'region name'],
    'village_taluka': ['village/taluka', 'village / taluka', 'village_taluka', 'village and taluka', 'village / taluka name', 'village taluka'],
    'completed_hydro_electric_projects': ['completed hydro electric projects', 'completed hydro projects', 'hydro electric projects', 'project name', 'completed_hydro_electric_projects'],
    'department': ['department', 'dept', 'dept name', 'department name'],
    'division': ['division', 'div', 'division name'],
    'consumer_no': ['consumer no', 'consumer no.', 'consumer_no', 'consumer number', 'consumer id', 'consumer_number'],
    'building_name': ['building name', 'building_name', 'name of building'],
    'address': ['address', 'building address', 'site address'],
    'pin_code': ['pin code', 'pincode', 'pin_code', 'pin', 'postal code'],
    'ownership_type': ['ownership type', 'ownership_type', 'ownership', 'type of ownership']
}


import re

def normalize_header_str(val):
    if val is None:
        return ''
    s = str(val).replace('\xa0', ' ').replace('\n', ' ').replace('\r', ' ')
    s = re.sub(r'[\/\\_\-\.\(\)]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip().lower()
    return s


def map_excel_columns(row, expected_headers, field_mapping):
    """
    Scans the given row as headers and maps each expected header to the correct Excel column index.
    Accurately handles ANY column order in the uploaded Excel file.
    """
    if not row:
        return {}

    raw_cells = [str(cell).strip() if cell is not None else '' for cell in row]
    norm_cells = [normalize_header_str(c) for c in raw_cells]

    mapped_headers = {}
    used_col_indices = set()

    # Pass 1: Exact normalized match with expected header name OR field_name
    for expected in expected_headers:
        field_name = field_mapping[expected]
        norm_expected = normalize_header_str(expected)
        norm_field = normalize_header_str(field_name)

        for col_idx, norm_cell in enumerate(norm_cells):
            if col_idx in used_col_indices or not norm_cell:
                continue
            if norm_cell == norm_expected or norm_cell == norm_field:
                mapped_headers[expected] = col_idx
                used_col_indices.add(col_idx)
                break

    # Pass 2: Exact match with defined aliases in HEADER_ALIASES
    for expected in expected_headers:
        if expected in mapped_headers:
            continue
        field_name = field_mapping[expected]
        aliases = [normalize_header_str(a) for a in HEADER_ALIASES.get(field_name, []) if a]

        for col_idx, norm_cell in enumerate(norm_cells):
            if col_idx in used_col_indices or not norm_cell:
                continue
            if norm_cell in aliases:
                mapped_headers[expected] = col_idx
                used_col_indices.add(col_idx)
                break

    # Pass 3: High-confidence substring match (e.g. "installed capacity (mw) bio")
    for expected in expected_headers:
        if expected in mapped_headers:
            continue
        field_name = field_mapping[expected]
        norm_expected = normalize_header_str(expected)
        aliases = [normalize_header_str(a) for a in HEADER_ALIASES.get(field_name, []) if a]

        for col_idx, norm_cell in enumerate(norm_cells):
            if col_idx in used_col_indices or not norm_cell or len(norm_cell) < 3:
                continue
            
            # Avoid confusing capacity_kw vs capacity_mw during substring match!
            if field_name == 'capacity_kw' and 'mw' in norm_cell:
                continue
            if field_name == 'capacity_mw' and 'kw' in norm_cell and 'mw' not in norm_cell:
                continue

            matched = False
            if len(norm_expected) >= 4 and (norm_expected in norm_cell or norm_cell in norm_expected):
                matched = True
            else:
                for alias in aliases:
                    if len(alias) >= 4 and (alias in norm_cell or norm_cell in alias):
                        matched = True
                        break

            if matched:
                mapped_headers[expected] = col_idx
                used_col_indices.add(col_idx)
                break

    return mapped_headers


def parse_date(date_val):
    """
    Utility function to safely parse Excel cell values into a Python datetime.date object.
    """
    if not date_val:
        return None
    if isinstance(date_val, datetime.date):
        return date_val
    if isinstance(date_val, datetime.datetime):
        return date_val.date()
    
    date_str = str(date_val).strip()
    if not date_str or date_str.lower() in ('none', 'null', 'n/a', '-', ''):
        return None
        
    # Check if numeric string representing Excel serial date (e.g., 44927 -> 2023-01-01)
    try:
        if date_str.isdigit() or (date_str.replace('.', '', 1).isdigit() and float(date_str) > 30000):
            num_days = float(date_str)
            if 30000 < num_days < 75000:
                base_date = datetime.date(1899, 12, 30)
                return base_date + datetime.timedelta(days=int(num_days))
    except Exception:
        pass

    # Attempt typical date parsing
    formats = [
        '%Y-%m-%d',
        '%d-%m-%Y',
        '%d/%m/%Y',
        '%Y/%m/%d',
        '%d.%m.%Y',
        '%Y.%m.%d',
        '%d-%b-%Y',
        '%d-%B-%Y',
        '%d %b %Y',
        '%d %B %Y',
        '%b %d, %Y',
        '%B %d, %Y',
        '%Y-%m-%dT%H:%M:%S',
        '%Y-%m-%dT%H:%M:%S.%fZ',
    ]
    for fmt in formats:
        try:
            if 'T' in date_str and fmt == '%Y-%m-%d':
                return datetime.datetime.strptime(date_str.split('T')[0], '%Y-%m-%d').date()
            return datetime.datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def parse_decimal(val):
    """
    Utility function to safely parse Excel numerical entries into Python Decimal,
    extracting numbers even if units (e.g. '10 MW') or formatting are present.
    """
    if val is None:
        return None
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ('none', 'null', 'n/a', '-'):
        return None
    match = re.search(r'[-+]?\d*\.?\d+', val_str.replace(',', ''))
    if match:
        try:
            return Decimal(match.group(0))
        except Exception:
            return None
    return None


class DownloadTemplateView(APIView):
    """
    Generates and downloads an empty Excel template containing only the column headers.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, energy_type):
        energy_type = energy_type.lower().replace('-', '_')
        if energy_type not in CONFIG_MAP:
            return Response(
                {"error": f"Invalid or unsupported energy type '{energy_type}'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        config = CONFIG_MAP[energy_type]
        headers = config['headers']

        # Create Excel workbook in memory
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{config['display_name'][:30]} Template"

        # Write headers in the first row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = openpyxl.styles.Font(bold=True)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="{energy_type}_template.xlsx"'
        wb.save(response)
        return response


class DownloadFilledDataView(APIView):
    """
    Downloads an Excel sheet containing all currently stored data rows for the selected energy type.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, energy_type):
        energy_type = energy_type.lower().replace('-', '_')
        if energy_type not in CONFIG_MAP:
            return Response(
                {"error": f"Invalid or unsupported energy type '{energy_type}'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        config = CONFIG_MAP[energy_type]
        model = config['model']
        headers = config['headers']
        field_mapping = config['fields']

        # Query all records
        records = model.objects.all().order_by('id')

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{config['display_name'][:30]} Data"

        # Write bold headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = openpyxl.styles.Font(bold=True)

        # Write data rows
        for row_idx, record in enumerate(records, start=2):
            for col_idx, header in enumerate(headers, start=1):
                field_name = field_mapping[header]
                val = getattr(record, field_name)
                if isinstance(val, Decimal):
                    val = float(val)
                ws.cell(row=row_idx, column=col_idx, value=val)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="{energy_type}_data.xlsx"'
        wb.save(response)
        return response


def read_spreadsheet_rows(excel_file):
    """
    Reads rows from .xlsx, .xls, or .csv files seamlessly.
    Returns a list of tuples/lists of cell values.
    """
    if hasattr(excel_file, 'seek'):
        excel_file.seek(0)
    file_bytes = excel_file.read()
    if hasattr(excel_file, 'seek'):
        excel_file.seek(0)

    # Attempt 1: openpyxl (.xlsx)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            return rows
    except Exception:
        pass

    # Attempt 2: xlrd (.xls legacy Excel 97-2003)
    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
        rows = []
        for r in range(ws.nrows):
            row_vals = []
            for c in range(ws.ncols):
                cell_val = ws.cell_value(r, c)
                if ws.cell_type(r, c) == xlrd.XL_CELL_DATE:
                    try:
                        date_tuple = xlrd.xldate_as_tuple(cell_val, wb.datemode)
                        cell_val = datetime.date(date_tuple[0], date_tuple[1], date_tuple[2])
                    except Exception:
                        pass
                row_vals.append(cell_val)
            rows.append(row_vals)
        if rows:
            return rows
    except Exception:
        pass

    # Attempt 3: CSV file
    try:
        import csv
        text_content = file_bytes.decode('utf-8-sig', errors='ignore')
        reader = csv.reader(io.StringIO(text_content))
        rows = [list(r) for r in reader if any(cell.strip() for cell in r)]
        if rows:
            return rows
    except Exception:
        pass

    raise ValueError("Could not read spreadsheet file. Please upload a valid Excel (.xlsx / .xls) or CSV file.")


class UploadExcelView(APIView):
    """
    Accepts an uploaded Excel file, parses it, and inserts/updates records for the selected energy type.
    Handles ANY column order in the uploaded Excel file.
    """
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, energy_type):
        energy_type = energy_type.lower().replace('-', '_')
        if energy_type not in CONFIG_MAP:
            return Response(
                {"error": f"Invalid or unsupported energy type '{energy_type}'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response(
                {"error": "No file uploaded. Please upload a valid Excel (.xlsx / .xls) file."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Open workbook with fallback for .xlsx, .xls, and .csv
        try:
            rows = read_spreadsheet_rows(excel_file)
        except Exception as e:
            return Response(
                {"error": f"Failed to read the Excel file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not rows:
            return Response(
                {"error": "The uploaded Excel sheet is empty."},
                status=status.HTTP_400_BAD_REQUEST
            )

        config = CONFIG_MAP[energy_type]
        model = config['model']
        field_mapping = config['fields']
        expected_headers = config['headers']

        # Dynamically scan top rows (up to row 15) to locate header row and map columns accurately
        best_header_row_idx = 0
        best_mapped_headers = {}
        max_matches = 0
        detected_file_headers = []

        for r_idx in range(min(15, len(rows))):
            row = rows[r_idx]
            if not row:
                continue

            current_mapped_headers = map_excel_columns(row, expected_headers, field_mapping)

            if len(current_mapped_headers) > max_matches:
                max_matches = len(current_mapped_headers)
                best_mapped_headers = current_mapped_headers
                best_header_row_idx = r_idx
                detected_file_headers = [str(c).strip() for c in row if c is not None and str(c).strip()]

        header_to_col_idx = best_mapped_headers

        # Positional Fallback: ONLY if no headers match by string, map columns by position 0, 1, 2...
        if not header_to_col_idx and rows:
            first_row = rows[0]
            if first_row:
                best_header_row_idx = 0
                for idx, expected in enumerate(expected_headers):
                    if idx < len(first_row):
                        header_to_col_idx[expected] = idx
                detected_file_headers = [str(c).strip() for c in first_row if c is not None]

        if not header_to_col_idx:
            return Response(
                {
                    "error": "No recognized columns found in the uploaded Excel file.",
                    "expected_columns": expected_headers,
                    "provided_columns": detected_file_headers or [str(c) for c in (rows[0] if rows else [])]
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        success_count = 0
        failed_rows = []

        # Process each data row starting after detected header row
        start_row_idx = best_header_row_idx + 1 if max_matches > 0 else 1
        for r_idx, row in enumerate(rows[start_row_idx:], start=start_row_idx + 1):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            record_data = {}
            has_some_data = False
            try:
                for header, col_idx in header_to_col_idx.items():
                    field_name = field_mapping[header]
                    val = row[col_idx] if col_idx < len(row) else None
                    
                    model_field = model._meta.get_field(field_name)
                    
                    if isinstance(model_field, models.DecimalField):
                        parsed_val = parse_decimal(val)
                        record_data[field_name] = parsed_val
                        if parsed_val is not None:
                            has_some_data = True
                    elif isinstance(model_field, models.DateField):
                        parsed_val = parse_date(val)
                        record_data[field_name] = parsed_val
                        if parsed_val is not None:
                            has_some_data = True
                    else:
                        if isinstance(val, float) and val.is_integer():
                            str_val = str(int(val)).strip()
                        elif val is not None:
                            str_val = str(val).strip()
                        else:
                            str_val = ''
                        if isinstance(model_field, models.CharField) and model_field.max_length:
                            str_val = str_val[:model_field.max_length]
                        record_data[field_name] = str_val
                        if str_val:
                            has_some_data = True

                # Only save record if row contained at least one non-empty value
                if has_some_data:
                    model.objects.create(**record_data)
                    success_count += 1
            except Exception as e:
                failed_rows.append({
                    "row_number": r_idx,
                    "error": str(e),
                    "data": {k: str(v) for k, v in record_data.items()}
                })

        if success_count == 0:
            error_msg = "0 records were saved into the database."
            if failed_rows:
                error_msg += f" {len(failed_rows)} row(s) failed. Error on row {failed_rows[0]['row_number']}: {failed_rows[0]['error']}"
            else:
                error_msg += " Please check if your uploaded file contains data rows beneath the column header row."

            return Response({
                "success": False,
                "error": error_msg,
                "imported_count": 0,
                "failed_count": len(failed_rows),
                "failed_rows": failed_rows
            }, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            "success": True,
            "message": f"Excel import completed: {success_count} records imported successfully." + (f" ({len(failed_rows)} rows failed)." if failed_rows else ""),
            "imported_count": success_count,
            "failed_count": len(failed_rows),
            "failed_rows": failed_rows
        }, status=status.HTTP_200_OK)


class EnergyDataListView(APIView):
    """
    Returns stored database rows in JSON format for the selected energy type.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, energy_type):
        energy_type = energy_type.lower().replace('-', '_')
        if energy_type not in CONFIG_MAP:
            return Response(
                {"error": f"Invalid or unsupported energy type '{energy_type}'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        config = CONFIG_MAP[energy_type]
        model = config['model']
        fields_to_fetch = list(set(['id'] + list(config['fields'].values())))
        model_field_names = set(f.name for f in model._meta.fields)
        valid_fields = [f for f in fields_to_fetch if f in model_field_names]
        records = list(model.objects.all().order_by('id').values(*valid_fields))

        return Response({
            "success": True,
            "display_name": config['display_name'],
            "headers": config['headers'],
            "fields": config['fields'],
            "data": records
        }, status=status.HTTP_200_OK)


class EnergyAnalyticsView(APIView):
    """
    Returns aggregated analytics for the specified energy type (e.g. government building solarization).
    Provides total projects (count of rows), total installed capacity MW,
    and district-wise distribution with counts, capacity, and percentage.
    """
    permission_classes = [permissions.AllowAny]

    def get(self, request, energy_type):
        raw_key = energy_type.lower().replace('-', '_')
        ALIASES = {
            'municipal_waste': 'msw',
            'municipal_solid_waste': 'msw',
            'small_hydro': 'shp',
            'small_hydro_power': 'shp',
            'small_hydro_projects': 'shp',
            'solar_kusum': 'solar_kusum',
            'kusum': 'solar_kusum',
            'kusum_ac': 'solar_kusum',
            'pm_kusum': 'solar_kusum',
            'solar_grid_conn': 'solar_grid',
            'solar_grid': 'solar_grid',
            'government_building_solarization': 'govt_solarization',
            'govt_building_solar': 'govt_solarization',
            'mskvy_2': 'mskvy',
            'mskvy_2_0': 'mskvy',
            'solar_mskvy': 'mskvy',
        }
        resolved_type = ALIASES.get(raw_key, raw_key)
        if resolved_type not in CONFIG_MAP:
            return Response(
                {"error": f"Invalid or unsupported energy type '{energy_type}'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        config = CONFIG_MAP[resolved_type]
        model = config['model']

        total_projects = model.objects.count()
        
        mw_field = next((f.name for f in model._meta.fields if f.name in ('capacity_mw', 'commissioned_capacity_mw', 'installed_capacity_mw')), None)
        kw_field = next((f.name for f in model._meta.fields if f.name in ('capacity_kw', 'commissioned_capacity_kw')), None)
        date_field = next((f.name for f in model._meta.fields if f.name in ('standardized_date', 'date_of_commissioned', 'commissioned_date', 'commission_date', 'date_of_commissioning')), None)
        year_field = next((f.name for f in model._meta.fields if f.name in ('comissioning_year', 'commissioning_year', 'year')), None)
        has_district = any(f.name == 'district' for f in model._meta.fields)
        has_division = any(f.name == 'division' for f in model._meta.fields)
        has_type = any(f.name == 'type' for f in model._meta.fields)

        agg = {}
        if mw_field:
            agg['total_mw'] = models.Sum(mw_field)
        if kw_field:
            agg['total_kw'] = models.Sum(kw_field)

        agg_res = model.objects.aggregate(**agg) if agg else {}
        total_mw = float(agg_res.get('total_mw') or 0.0)
        total_kw = float(agg_res.get('total_kw') or 0.0)

        # Type breakdown (e.g. for Solar Kusum: KUSUM A vs KUSUM C)
        type_breakdown = {}
        if has_type and mw_field:
            for t_item in model.objects.values('type').annotate(mw=models.Sum(mw_field), count=models.Count('id')):
                t_label = str(t_item.get('type') or 'General').strip()
                type_breakdown[t_label] = {
                    'type': t_label,
                    'capacity_mw': round(float(t_item['mw'] or 0.0), 2),
                    'count': t_item['count']
                }

        # Source breakdown (e.g. for MSKVY: MSKVY 1.0 vs MSKVY 2.0)
        source_breakdown = {}
        has_source = any(f.name == 'source' for f in model._meta.fields)
        if has_source and mw_field:
            for s_item in model.objects.values('source').annotate(mw=models.Sum(mw_field), count=models.Count('id')):
                s_label = str(s_item.get('source') or 'General').strip()
                source_breakdown[s_label] = {
                    'source': s_label,
                    'capacity_mw': round(float(s_item['mw'] or 0.0), 2),
                    'count': s_item['count']
                }

        district_data = []
        if has_district:
            raw_district_counts = {}
            query_fields = ['district']
            if mw_field:
                query_fields.append(mw_field)

            for r in model.objects.values(*query_fields):
                raw_dist = str(r.get('district') or '').strip()
                if not raw_dist:
                    continue
                d_norm = raw_dist.title().rstrip(' ,;.')
                d_lower = d_norm.lower()
                if 'sambhajinagar' in d_lower or 'aurangabad' in d_lower:
                    d_norm = 'Chhatrapati Sambhajinagar'
                elif 'ahmednagar' in d_lower or 'ahemadnagar' in d_lower or 'ahmadnagar' in d_lower or 'ahilyanagar' in d_lower or 'ahilya' in d_lower:
                    d_norm = 'Ahilyanagar'
                elif 'osmanabad' in d_lower or 'usmanabad' in d_lower or 'dharashiv' in d_lower:
                    d_norm = 'Dharashiv'
                elif d_lower in ('sangli', 'sangali'):
                    d_norm = 'Sangli'
                elif d_lower in ('amravati', 'amaravati'):
                    d_norm = 'Amravati'
                elif d_lower in ('nandurbar', 'nadurbar'):
                    d_norm = 'Nandurbar'
                
                if d_norm not in raw_district_counts:
                    raw_district_counts[d_norm] = {'count': 0, 'capacity_mw': 0.0}
                raw_district_counts[d_norm]['count'] += 1
                if mw_field and r.get(mw_field) is not None:
                    try:
                        raw_district_counts[d_norm]['capacity_mw'] += float(r[mw_field])
                    except (ValueError, TypeError):
                        pass

            # Sort by count for govt-solarization, or by capacity_mw for energy projects
            is_count_primary = resolved_type in ('govt_solarization', 'government_building_solarization')
            sorted_districts = sorted(
                raw_district_counts.items(),
                key=(lambda x: x[1]['count']) if is_count_primary else (lambda x: (x[1]['capacity_mw'], x[1]['count'])),
                reverse=True
            )

            palette = [
                "#2563eb", "#1e3a8a", "#ea580c", "#4f46e5", "#059669", 
                "#7c3aed", "#0d9488", "#e11d48", "#d97706", "#0891b2", 
                "#16a34a", "#c2410c", "#9333ea", "#3b82f6", "#64748b",
                "#b45309", "#06b6d4", "#84cc16", "#be185d", "#4338ca"
            ]

            for idx, (d_name, stats) in enumerate(sorted_districts):
                cnt = stats['count']
                mw_val = round(stats['capacity_mw'], 3)
                if not is_count_primary and total_mw > 0:
                    pct = round((mw_val / total_mw * 100), 2)
                else:
                    pct = round((cnt / total_projects * 100), 2) if total_projects > 0 else 0.0
                
                if mw_val >= 1000:
                    formatted_k = f"{mw_val/1000:.2f}K"
                elif mw_val >= 100:
                    formatted_k = f"{mw_val:.1f}"
                elif mw_val > 0:
                    formatted_k = f"{mw_val:.2f}"
                else:
                    formatted_k = f"{cnt}"

                district_data.append({
                    'district': d_name,
                    'count': cnt,
                    'capacity_mw': mw_val,
                    'percentage': pct,
                    'formatted_count': formatted_k,
                    'color': palette[idx % len(palette)]
                })

        # Timeline aggregation (Capacity over years strictly from database)
        timeline_data = []
        if (date_field or year_field) and mw_field:
            yearly_map = {}
            query_fields_tl = [mw_field]
            if date_field:
                query_fields_tl.append(date_field)
            if year_field and year_field != date_field:
                query_fields_tl.append(year_field)

            for r in model.objects.values(*query_fields_tl):
                d_val = r.get(date_field) if date_field else None
                yr = None
                if d_val and hasattr(d_val, 'year'):
                    yr = d_val.year
                elif d_val:
                    try:
                        yr = int(str(d_val)[:4])
                    except (ValueError, TypeError):
                        yr = None

                # Fallback to year_field if date_field was empty or out of bounds
                if (yr is None or not (1950 <= yr <= 2035)) and year_field:
                    y_val = r.get(year_field)
                    if y_val:
                        try:
                            match = re.search(r'\b(19\d\d|20\d\d)\b', str(y_val).strip())
                            if match:
                                yr = int(match.group(1))
                        except Exception:
                            yr = None

                if yr and 1950 <= yr <= 2035:
                    if yr not in yearly_map:
                        yearly_map[yr] = {'added_mw': 0.0, 'count': 0}
                    try:
                        yearly_map[yr]['added_mw'] += float(r.get(mw_field) or 0.0)
                        yearly_map[yr]['count'] += 1
                    except (ValueError, TypeError):
                        pass

            cum_mw = 0.0
            cum_cnt = 0
            for yr in sorted(yearly_map.keys()):
                cum_mw += yearly_map[yr]['added_mw']
                cum_cnt += yearly_map[yr]['count']
                timeline_data.append({
                    'year': yr,
                    'cumulative_mw': round(cum_mw, 2),
                    'added_mw': round(yearly_map[yr]['added_mw'], 2),
                    'cumulative_count': cum_cnt,
                    'added_count': yearly_map[yr]['count']
                })

        division_data = []
        if has_division:
            raw_div_counts = {}
            query_fields_div = ['division']
            if mw_field:
                query_fields_div.append(mw_field)
            
            for r in model.objects.values(*query_fields_div):
                raw_div = str(r.get('division') or '').strip()
                if not raw_div:
                    continue
                div_norm = raw_div.title()
                if 'chha' in div_norm.lower() or 'sambhajinagar' in div_norm.lower() or 'aurangabad' in div_norm.lower():
                    div_norm = 'Chhatrapati Sambhajinagar'
                
                if div_norm not in raw_div_counts:
                    raw_div_counts[div_norm] = {'count': 0, 'capacity_mw': 0.0}
                raw_div_counts[div_norm]['count'] += 1
                if mw_field and r.get(mw_field) is not None:
                    try:
                        raw_div_counts[div_norm]['capacity_mw'] += float(r[mw_field])
                    except (ValueError, TypeError):
                        pass

            sorted_divisions = sorted(
                raw_div_counts.items(),
                key=lambda x: x[1]['capacity_mw'],
                reverse=True
            )
            for d_name, stats in sorted_divisions:
                cnt = stats['count']
                mw_val = round(stats['capacity_mw'], 3)
                pct = round((mw_val / total_mw * 100), 2) if total_mw > 0 else 0.0
                division_data.append({
                    'division': d_name,
                    'count': cnt,
                    'capacity_mw': mw_val,
                    'percentage': pct
                })

        return Response({
            'success': True,
            'energy_type': energy_type,
            'display_name': config['display_name'],
            'total_projects': total_projects,
            'total_capacity_mw': round(total_mw, 2),
            'total_capacity_kw': round(total_kw, 2),
            'districts': district_data,
            'divisions': division_data,
            'timeline': timeline_data,
            'types': type_breakdown,
            'sources': source_breakdown
        }, status=status.HTTP_200_OK)

